from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from io import BytesIO

from pydantic import BaseModel
import json
import os
import random
from datetime import datetime, timedelta
from collections import Counter
from typing import List

app = FastAPI()

# ------------------ CORS ------------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ------------------ ARCHIVOS ESTÁTICOS ------------------
app.mount("/static", StaticFiles(directory="static"), name="static")


# ------------------ CARGA DE USUARIOS ------------------
json_path = os.path.join(os.path.dirname(__file__), "users.json")
inventario_path = os.path.join(os.path.dirname(__file__), "data", "inventario.json")

def load_users():
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print("⚠️ Error al cargar users.json:", e)
            return {}
    else:
        print("⚠️ Archivo users.json no encontrado.")
        return {}

def save_users(data):
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
def load_inventario():
    if os.path.exists(inventario_path):
        try:
            with open(inventario_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print("⚠️ Error al cargar inventario.json:", e)
            return {"items": []}
    else:
        return {"items": []}

def save_inventario(data):
    os.makedirs(os.path.dirname(inventario_path), exist_ok=True)
    with open(inventario_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
fake_users_db = load_users()
# ------------------ RUTAS: AUTENTICACIÓN ------------------
class LoginRequest(BaseModel):
    email: str
    password: str
    remember: bool

@app.post("/api/auth/login")
async def login_api(data: LoginRequest):
    try:
        with open("users.json", "r", encoding="utf-8") as f:
            users = json.load(f)
    except Exception:
        return JSONResponse(status_code=500, content={"detail": "Error al leer usuarios"})

    user = next(
        (u for u in users.values()
         if u.get("email") == data.email or u.get("username") == data.email),
        None
    )

    if not user or user.get("password") != data.password:
        return JSONResponse(status_code=401, content={"detail": "Credenciales inválidas"})

    return {
        "access_token": "fake-token-123",
        "user": {
            "email": user.get("email"),
            "username": user.get("username"),
            "name": user.get("name"),
            "role": user.get("role")
        }
    }

# ------------------ RECUPERACIÓN DE CONTRASEÑA ------------------
@app.post("/api/auth/forgot-password")
async def forgot_password(request: Request):
    datos = await request.json()
    email = datos.get("email")
    usuarios = load_users()

    usuario = next((u for u in usuarios.values() if u["email"] == email), None)
    if not usuario:
        return {"message": "Correo no registrado"}

    codigo = str(random.randint(100000, 999999))
    usuario["codigo_recuperacion"] = codigo
    save_users(usuarios)

    enviar_codigo(email, codigo)
    return {"message": "Código enviado al correo"}

# ------------------ DATOS DE USUARIO ------------------
@app.get("/api/user")
def get_user(email: str):
    usuarios = load_users()
    usuario = next((u for u in usuarios.values() if u["email"] == email), None)
    if not usuario:
        return JSONResponse(status_code=404, content={"detail": "Usuario no encontrado"})
    return usuario
from fastapi import Body

class Empleado(BaseModel):
    email: str
    username: str
    password: str
    name: str
    role: str
    estado: str | None = "activo"
    codigo_recuperacion: str | None = None

@app.get("/api/empleados")
def get_empleados():
    usuarios = load_users()
    empleados = []
    for email, data in usuarios.items():
        data["email"] = email
        if "estado" not in data:
            data["estado"] = "activo"
        empleados.append(data)
    return empleados

@app.post("/api/empleados")
def crear_empleado(emp: Empleado):
    usuarios = load_users()
    if emp.email in usuarios:
        raise HTTPException(status_code=400, detail="Ya existe un usuario con ese email")
    usuarios[emp.email] = emp.dict()
    save_users(usuarios)
    return {"message": "Empleado creado"}

@app.put("/api/empleados/{email}")
def actualizar_empleado(email: str, emp: Empleado):
    usuarios = load_users()
    if email not in usuarios:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    usuarios[email] = emp.dict()
    save_users(usuarios)
    return {"message": "Empleado actualizado"}

@app.delete("/api/empleados/{email}")
def eliminar_empleado(email: str):
    usuarios = load_users()
    if email not in usuarios:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    del usuarios[email]
    save_users(usuarios)
    return {"message": "Empleado eliminado"}
@app.get("/partial/empleados")
def get_empleados_section(email: str = ""):
    usuarios = load_users()
    usuario = next((u for u in usuarios.values() if u["email"] == email), None)

    if not usuario or usuario.get("role") != "admin":
        return JSONResponse(status_code=403, content={"error": "Acceso denegado"})

    return FileResponse("static/sections/empleados.html")
# ------------------ CARGA DE VENTAS ------------------
ventas_path = os.path.join(os.path.dirname(__file__), "data", "ventas.json")
def load_ventas():
    if os.path.exists(ventas_path):
        try:
            with open(ventas_path, "r", encoding="utf-8") as f:
                raw_data = json.load(f)
                labels = [item["mes"] for item in raw_data]
                data = [item["monto"] for item in raw_data]
                return {
                    "ventasPorPeriodo": {
                        "meses": labels,
                        "ventas": data,
                        "pedidos": [round(monto / 300) for monto in data]
                    },
                    "ventasPorZona": {
                        "zonas": ["Norte", "Sur", "Este", "Oeste"],
                        "valores": [32000, 28000, 19000, 21000]
                    }
                }
        except Exception as e:
            print("⚠️ Error al cargar ventas.json:", e)
            return {}
    else:
        print("⚠️ Archivo ventas.json no encontrado.")
        return {}

# ------------------ FUNCIÓN DE CORREO (SIMULADA) ------------------
def enviar_codigo(destinatario, codigo):
    print(f"📧 Simulación: se enviaría el código {codigo} al correo {destinatario}")


# ##################################################################
# ###           NUEVO: MODELO Y HELPERS PARA CLIENTES            ###
# ##################################################################

CLIENTES_JSON_PATH = os.path.join(os.path.dirname(__file__), "data", "clientes.json")

# --- MODELO PYDANTIC PARA UN CLIENTE ---

class Cliente(BaseModel):
    id: int | None = None
    nombre: str
    email: str | None = None  # Hacer opcionales los campos no obligatorios
    telefono: str | None = None
    zona: str | None = None
    direccion: str | None = None
    bidones: int | None = 0
    deuda: float | None = 0.0
    estado: str | None = "activo"

# --- FUNCIONES AUXILIARES PARA LEER/ESCRIBIR CLIENTES ---
def load_clientes():
    if not os.path.exists(CLIENTES_JSON_PATH):
        return [] # Importante: una lista vacía
    with open(CLIENTES_JSON_PATH, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return [] # Lista vacía si el JSON está corrupto

def save_clientes(data: list):
    with open(CLIENTES_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


# ------------------ RUTAS: PÁGINAS ------------------

# Página principal (login)
@app.get("/")
def read_root():
    return FileResponse("static/index.html")

# Página principal del panel (antes era dashboard)
@app.get("/base")
def base_panel():
    return FileResponse("static/base.html")

# Cargar secciones dinámicas dentro del panel (Dashboard, Clientes, etc.)
@app.get("/partial/{page}")
async def get_partial(page: str):
    file_path = os.path.join("static", "sections", f"{page}.html")
    if os.path.exists(file_path):
        return FileResponse(file_path)
    return JSONResponse(status_code=404, content={"error": "Sección no encontrada"})


# ==================== INVENTARIO ====================

# Obtener todos los productos del inventario
@app.get("/api/inventario")
def get_inventario():
    inventario = load_inventario()
    return inventario

# Obtener un producto específico por ID
@app.get("/api/inventario/{producto_id}")
def get_producto(producto_id: int):
    inventario = load_inventario()
    producto = next((p for p in inventario["items"] if p["id"] == producto_id), None)
    if not producto:
        return JSONResponse(status_code=404, content={"error": "Producto no encontrado"})
    return producto

# Crear nuevo producto
class ProductoRequest(BaseModel):
    nombre: str
    stock: int
    precio: float
    categoria: str = "General"
    descripcion: str = ""

@app.post("/api/inventario")
async def crear_producto(data: ProductoRequest):
    inventario = load_inventario()
    
    # Generar ID único
    nuevo_id = max([p.get("id", 0) for p in inventario["items"]], default=0) + 1
    
    nuevo_producto = {
        "id": nuevo_id,
        "nombre": data.nombre,
        "stock": data.stock,
        "precio": data.precio,
        "categoria": data.categoria,
        "descripcion": data.descripcion
    }
    
    inventario["items"].append(nuevo_producto)
    save_inventario(inventario)
    
    return {"message": "Producto creado exitosamente", "producto": nuevo_producto}

# Actualizar producto
@app.put("/api/inventario/{producto_id}")
async def actualizar_producto(producto_id: int, data: ProductoRequest):
    inventario = load_inventario()
    producto = next((p for p in inventario["items"] if p["id"] == producto_id), None)
    
    if not producto:
        return JSONResponse(status_code=404, content={"error": "Producto no encontrado"})
    
    producto["nombre"] = data.nombre
    producto["stock"] = data.stock
    producto["precio"] = data.precio
    producto["categoria"] = data.categoria
    producto["descripcion"] = data.descripcion
    
    save_inventario(inventario)
    
    return {"message": "Producto actualizado exitosamente", "producto": producto}

# Eliminar producto
@app.delete("/api/inventario/{producto_id}")
async def eliminar_producto(producto_id: int):
    inventario = load_inventario()
    inventario["items"] = [p for p in inventario["items"] if p["id"] != producto_id]
    
    save_inventario(inventario)
    
    return {"message": "Producto eliminado exitosamente"}

# Obtener estadísticas del inventario
@app.get("/api/inventario-stats")
def get_inventario_stats():
    inventario = load_inventario()
    items = inventario["items"]
    
    total_stock = sum(p.get("stock", 0) for p in items)
    valor_total = sum(p.get("stock", 0) * p.get("precio", 0) for p in items)
    cantidad_productos = len(items)
    productos_bajo_stock = sum(1 for p in items if p.get("stock", 0) < 20)
    
    return {
        "total_stock": total_stock,
        "valor_total": valor_total,
        "cantidad_productos": cantidad_productos,
        "productos_bajo_stock": productos_bajo_stock
    }

# ------------------ RUTAS: DASHBOARD Y REPORTES ------------------

@app.get("/api/dashboard/ventas")
def get_ventas():
    return load_ventas()

@app.get("/api/dashboard/entregas")
def get_entregas():
    return {
        "labels": ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"],
        "data": [45, 52, 48, 62, 55, 68, 42]
    }
@app.get("/api/reportes")
def get_reportes(periodo: int = 6, zona: str = "todas"):
    """Retorna datos para los gráficos de reportes con filtros"""
    try:
        with open("data/inventario.json") as f:
            inventario = json.load(f).get("items", [])

        with open("data/clientes.json") as f:
            clientes = json.load(f)

        with open("data/ventas.json") as f:
            ventas = json.load(f)

        print(f"📊 Total ventas antes de filtrar: {len(ventas)}")

        # Filtrar por zona
        if zona != "todas":
            ventas = [v for v in ventas if v.get("zona", "").lower() == zona.lower()]
            print(f"📊 Ventas después de filtrar por zona '{zona}': {len(ventas)}")

        hoy = datetime.now()
        meses_labels = []
        meses_dict = {}

        for i in range(periodo - 1, -1, -1):
            fecha = hoy - timedelta(days=30 * i)
            mes_key = fecha.strftime("%Y-%m")
            mes_label = fecha.strftime("%b")
            meses_labels.append(mes_label)
            meses_dict[mes_key] = 0

        # Ventas mensuales
        ventasMensuales = []
        for mes_key in sorted(meses_dict.keys()):
            total = sum(
                v.get("total_calculado", v.get("monto", 0))
                for v in ventas
                if v.get("fecha", "").startswith(mes_key)
            )
            ventasMensuales.append(total)

        # Ventas por zona
        zonas_dict = {}
        for v in ventas:
            zona_actual = v.get("zona", "Sin zona").capitalize()
            monto = v.get("total_calculado", v.get("monto", 0))
            zonas_dict[zona_actual] = zonas_dict.get(zona_actual, 0) + monto

        ventasPorZona = [
            {"zona": z, "ventas": round(v, 2)} for z, v in zonas_dict.items()
        ]

        # Diccionario id → nombre real del producto
        id_to_nombre = {
            p["id"]: p["nombre"]
            for p in inventario
            if "id" in p and "nombre" in p
        }

        # Productos top
        productos_vendidos = Counter()
        for venta in ventas:
            for p in venta.get("productos", []):
                productos_vendidos[p["producto_id"]] += p.get("cantidad", 1)

        productosTop = [
            {"nombre": id_to_nombre.get(pid, f"Producto {pid}"), "ventas": count}
            for pid, count in productos_vendidos.most_common(5)
        ]

        # Clientes top
        clientes_compras = Counter()
        for venta in ventas:
            if "cliente_id" in venta:
                clientes_compras[venta["cliente_id"]] += venta.get("monto", 0)

        clientesTop = [
            {
                "nombre": next(
                    (c["nombre"] for c in clientes if c["id"] == cid),
                    f"Cliente {cid}",
                ),
                "compras": monto,
            }
            for cid, monto in clientes_compras.most_common(5)
        ]
        clientesDeudores = sorted(
            [c for c in clientes if c.get("deuda", 0) > 0],
            key=lambda c: c["deuda"],
            reverse=True
        )[:5]

        clientesDeudores = [
            {"nombre": c["nombre"], "deuda": round(c["deuda"], 2)}
            for c in clientesDeudores
        ]



        

        return {
            "meses": meses_labels,
            "ventasMensuales": ventasMensuales,
            "ventasPorZona": ventasPorZona,
            "productosTop": productosTop,
            "clientesTop": clientesTop,
            "clientesDeudores": clientesDeudores,
        }

    except Exception as e:
        print(f"❌ Error en /api/reportes: {str(e)}")
        return {
            "error": str(e),
            "meses": [],
            "ventasMensuales": [],
            "ventasPorZona": [],
            "productosTop": [],
            "clientesTop": [],
            "clientesDeudores": [],
        }
    
    
@app.get("/api/reportes/exportar")
def exportar_reportes(periodo: int = 6, zona: str = "todas"):
    datos = get_reportes(periodo, zona)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Ventas Mensuales"
    ws1.append(["Mes", "Monto"])
    for mes, monto in zip(datos["meses"], datos["ventasMensuales"]):
        ws1.append([mes, monto])

    ws2 = wb.create_sheet("Ventas por Zona")
    ws2.append(["Zona", "Monto"])
    for z in datos["ventasPorZona"]:
        ws2.append([z["zona"], z["ventas"]])

    ws3 = wb.create_sheet("Productos Top")
    ws3.append(["Producto", "Ventas"])
    for p in datos["productosTop"]:
        ws3.append([p["nombre"], p["ventas"]])

    ws4 = wb.create_sheet("Clientes Top")
    ws4.append(["Cliente", "Compras"])
    for c in datos["clientesTop"]:
        ws4.append([c["nombre"], c["compras"]])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reportes.xlsx"}
    )
# ##################################################################
# ###           NUEVO: ENDPOINTS DE API PARA CLIENTES            ###
# ##################################################################

@app.get("/api/clientes")
def get_todos_los_clientes():
    """
    Endpoint para OBTENER (GET) todos los clientes.
    """
    return load_clientes()

@app.post("/api/clientes")
def agregar_nuevo_cliente(cliente: Cliente):
    """
    Endpoint para AÑADIR (POST) un nuevo cliente.
    'cliente: Cliente' usa el modelo Pydantic para validar los datos.
    """
    clientes = load_clientes()
    
    # Generamos un ID nuevo (basado en el ID más alto que exista)
    nuevo_id = max([c.get("id", 0) for c in clientes] + [0]) + 1
    
    # Convertimos el modelo de Pydantic a un diccionario
    nuevo_cliente_dict = cliente.model_dump()
    nuevo_cliente_dict["id"] = nuevo_id
    
    # Añadimos el nuevo cliente a la lista
    clientes.append(nuevo_cliente_dict)
    
    # Guardamos la lista actualizada en el archivo JSON
    save_clientes(clientes)
    
    print(f"✅ Cliente agregado: {nuevo_cliente_dict['nombre']}")
    return nuevo_cliente_dict # Devolvemos el cliente creado (con su ID)

@app.delete("/api/clientes/{cliente_id}")
def eliminar_cliente(cliente_id: int):
    """
    Endpoint para BORRAR (DELETE) un cliente por su ID.
    """
    clientes = load_clientes()
    
    # Filtramos la lista, quedándonos con todos MENOS el que queremos borrar
    clientes_filtrados = [c for c in clientes if c.get("id") != cliente_id]
    
    # Si las listas tienen el mismo tamaño, es que no se encontró el ID
    if len(clientes) == len(clientes_filtrados):
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    # Guardamos la lista filtrada
    save_clientes(clientes_filtrados)
    
    print(f"🗑️ Cliente eliminado con ID: {cliente_id}")
    return {"mensaje": f"Cliente con ID {cliente_id} eliminado"}
#Modificar cliente
@app.put("/api/clientes/{cliente_id}")
def actualizar_cliente(cliente_id: int, cliente_actualizado: Cliente):
    """
    Endpoint para ACTUALIZAR (PUT) un cliente existente.
    """
    clientes = load_clientes()
    
    # Buscamos el índice del cliente a actualizar
    index_a_actualizar = -1
    for i, cliente in enumerate(clientes):
        if cliente.get("id") == cliente_id:
            index_a_actualizar = i
            break
            
    # Si no lo encontramos, error 404
    if index_a_actualizar == -1:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
        
    # Convertimos el modelo Pydantic a un diccionario
    datos_actualizados = cliente_actualizado.model_dump()
    
    # MUY IMPORTANTE: Nos aseguramos de mantener el ID original
    datos_actualizados["id"] = cliente_id 
    
    # Reemplazamos el cliente viejo con los datos actualizados
    clientes[index_a_actualizar] = datos_actualizados
    
    # Guardamos la lista completa en el JSON
    save_clientes(clientes)
    
    print(f"🔄 Cliente actualizado con ID: {cliente_id}")
    return datos_actualizados # Devolvemos el cliente actualizado

# ------------------ MONTAR DIRECTORIO DE DATOS ------------------
# (Debe ir al final)
# ##################################################################
# ###           NUEVO: MODELO Y ENDPOINTS PARA VENTAS            ###
# ##################################################################

VENTAS_JSON_PATH = os.path.join(os.path.dirname(__file__), "data", "ventas.json")

# --- MODELO PYDANTIC PARA VENTA ---

class ProductoVendido(BaseModel):
    producto_id: int
    cantidad: int

class Venta(BaseModel):
    cliente_id: int
    monto: float
    tipo_pago: str
    productos: List[ProductoVendido] = []


# --- FUNCIONES AUXILIARES PARA LEER/ESCRIBIR VENTAS ---
def load_ventas_list():
    """Carga la lista de ventas (array) para la tabla"""
    if not os.path.exists(VENTAS_JSON_PATH):
        return []
    with open(VENTAS_JSON_PATH, "r", encoding="utf-8") as f:
        try:
            data = json.load(f)
            # Si por error el archivo tiene el formato viejo de gráficos (diccionario), devolvemos lista vacía
            if isinstance(data, dict): 
                return [] 
            return data
        except json.JSONDecodeError:
            return []

def save_ventas_list(data: list):
    with open(VENTAS_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

# --- ENDPOINT 1: OBTENER HISTORIAL (GET) ---
@app.get("/api/ventas")
def get_historial_ventas():
    """Devuelve la lista cruda de ventas para la tabla"""
    ventas = load_ventas_list()
    # Ordenamos por fecha descendente (la más nueva primero)
    return sorted(ventas, key=lambda v: v.get("fecha", ""), reverse=True)

# --- ENDPOINT 2: REGISTRAR VENTA (POST) ---
@app.post("/api/ventas")
async def registrar_venta(venta: Venta, request: Request):
    clientes = load_clientes()
    inventario = load_inventario()
    ventas = load_ventas_list()

    extra = await request.json()
    body_raw = venta.model_dump()
    nueva_venta_dict = {**body_raw, **extra}

    cliente_actualizado = None
    cliente_nombre = "Desconocido"
    cliente_zona = "Sin zona"

    for i, c in enumerate(clientes):
        if c.get("id") == venta.cliente_id:
            cliente_actualizado = c
            cliente_nombre = c.get("nombre", "Desconocido")
            cliente_zona = c.get("zona", "Sin zona")
            break

    if not cliente_actualizado:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    # Validar stock antes de descontar
    if not nueva_venta_dict.get("is_pago_de_deuda"):
        for item in venta.productos:
            producto = next((p for p in inventario["items"] if p["id"] == item.producto_id), None)
            if not producto:
                raise HTTPException(status_code=404, detail=f"Producto ID {item.producto_id} no encontrado")
            if producto["stock"] < item.cantidad:
                raise HTTPException(status_code=400, detail=f"Stock insuficiente para {producto['nombre']} (disponible: {producto['stock']}, solicitado: {item.cantidad})")

    # Calcular total
    total_calculado = 0
    for item in venta.productos:
        producto = next((p for p in inventario["items"] if p["id"] == item.producto_id), None)
        total_calculado += producto.get("precio", 0) * item.cantidad

    monto_pagado = venta.monto
    deuda_real = max(0, total_calculado - monto_pagado)

    # Actualizar deuda del cliente
    if nueva_venta_dict.get("tipo_pago") == "fiado":
        cliente_actualizado["deuda"] = cliente_actualizado.get("deuda", 0) + deuda_real
    elif nueva_venta_dict.get("tipo_pago") == "pago_de_deuda":
        cliente_actualizado["deuda"] = max(0, cliente_actualizado.get("deuda", 0) - nueva_venta_dict.get("pago_cliente", 0))

    for i, c in enumerate(clientes):
        if c.get("id") == venta.cliente_id:
            clientes[i] = cliente_actualizado
            break

    # Descontar stock si corresponde
    if not nueva_venta_dict.get("is_pago_de_deuda"):
        for item in venta.productos:
            producto = next((p for p in inventario["items"] if p["id"] == item.producto_id), None)
            producto["stock"] -= item.cantidad

    save_clientes(clientes)
    save_inventario(inventario)

    nueva_venta_dict["id"] = int(datetime.now().timestamp())
    nueva_venta_dict["fecha"] = datetime.now().isoformat()
    nueva_venta_dict["cliente_nombre"] = cliente_nombre
    nueva_venta_dict["zona"] = cliente_zona
    nueva_venta_dict["total_calculado"] = total_calculado
    nueva_venta_dict["pago_cliente"] = monto_pagado
    nueva_venta_dict["deuda_generada"] = deuda_real

    ventas.append(nueva_venta_dict)
    save_ventas_list(ventas)

    print(f"📦 Venta registrada: Total={total_calculado} Pagado={monto_pagado} Deuda={deuda_real}")
    return {"mensaje": "Venta registrada", "venta_creada": nueva_venta_dict}

@app.delete("/api/ventas/{venta_id}")
def anular_venta(venta_id: int):
    """
    Anula una venta. 
    Si la venta era FIADA, se RESTA la deuda al cliente correspondiente.
    """
    ventas = load_ventas_list()
    clientes = load_clientes()
    
    # 1. Buscar la venta a eliminar
    venta_a_borrar = None
    index_venta = -1
    
    for i, v in enumerate(ventas):
        if v.get("id") == venta_id:
            venta_a_borrar = v
            index_venta = i
            break
    
    if not venta_a_borrar:
        raise HTTPException(status_code=404, detail="Venta no encontrada")

    # 2. Si fue FIADA, revertir la deuda del cliente
    if venta_a_borrar.get("tipo_pago") == "fiado":
        cliente_id = venta_a_borrar.get("cliente_id")
        monto = venta_a_borrar.get("monto", 0)
        
        # Buscar al cliente
        for i, c in enumerate(clientes):
            if c.get("id") == cliente_id:
                deuda_actual = c.get("deuda", 0)
                # RESTAMOS el monto (porque se anula la deuda)
                c["deuda"] = max(0, deuda_actual - monto) 
                clientes[i] = c
                save_clientes(clientes) # Guardar cambios en clientes
                print(f"📉 Deuda revertida para cliente {cliente_id}. -${monto}")
                break

    # 3. Eliminar la venta de la lista y guardar
    ventas.pop(index_venta)
    save_ventas_list(ventas)
    
    return {"mensaje": "Venta anulada correctamente"}
app.mount("/data", StaticFiles(directory="data"), name="data")